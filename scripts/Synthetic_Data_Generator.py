# -*- coding: utf-8 -*-
"""

@author: jon0003
"""

import os
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import lognorm, norm
import seaborn as sns
from math import exp

# set relative path to files
cwd = os.getcwd()
inputPath = os.path.join(cwd, 'Input Data')
outputPath = os.path.join(cwd, 'Output Data')
BOLRPath = (r'C:\Users\jon0003\OneDrive - Auburn University\CAREER\Research Tasks\7. Theory Guided Data Science\2_Analysis\B-OLR\Input Data')

data = pd.read_csv(os.path.join(inputPath, r'WiSPDv2.2.csv'))
fragilities = pd.read_csv(os.path.join(inputPath, r'fragility_fitted.csv'))     # These are the updated fragility functions from R2D(HAZUS5.1)
building_codes = load_workbook(filename = os.path.join(inputPath, 'BuildingCodeHistory_Municipality.xlsx'))

# convert wind speed to mph
data['wind_speed_mph'] = data['wind_speed_ms']* 2.23694


# Global variables
NUM_LIMIT_STATES = 4  # number of limit states
ROOF_COVER_ADJUSTMENTS = {
    'Metal': {
        (1, 2): 1.115,  # LS1 and LS2
        (3,): 1.115,    # LS3
    }
}
STORM_TYPE_ADJUSTMENTS = {
    'Tornado': {
        (1, 2, 3): 0.9129,  # LS1, LS2, and LS3
        (4,): 0.9535        # LS4
    }
}

terrain_col = 'z0_min'

# File names
WiSPD_filename = 'UpdatedFragilities_WiSPD_InferedHAZUSData_Weights_min_z0.csv'
HAZUS_filename = 'UpdatedFragilities_HAZUS_SyntheticData_Weights_min_z0.csv'

####################################################################################################
def compile_unique_codes(excel_file_path):
    # Load the Excel workbook
    workbook = pd.ExcelFile(excel_file_path)
    
    # Set to store unique codes
    unique_codes_set = set()
    
    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheet_names:
        # Read the current sheet into a DataFrame
        df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        
        # Check if 'Code' column exists
        if 'Code' in df.columns:
            # Extract the unique values from the 'Code' column and add to the set
            unique_codes_set.update(df['Code'].unique())
    
    # Convert the set to a DataFrame
    unique_codes_df = pd.DataFrame(list(unique_codes_set), columns=['Code'])
    
    return unique_codes_df

# Example usage
unique_codes_df = compile_unique_codes(os.path.join(inputPath, 'BuildingCodeHistory_Municipality.xlsx'))

# Edit WiSPD data 
data['wind_speed_mph'] = data['wind_speed_ms'] * 2.23694
data = data[data['wind_damage_rating'] != -1]
data = data[data['year_built'] != 0]
data = data[data['number_of_stories'] != 0]
# Drop rows where _project contains "Sally" - missing data
data = data[~data['_project'].str.contains('Sally', na=False)]
# Replace empty values with 0
data['fenestration_avg'] = data['fenestration_avg'].fillna(0)
data['roof_cover_damage'] = data['roof_cover_damage'].fillna(0)
data['roof_structure_damage'] = data['roof_structure_damage'].fillna(0)
data['roof_substrate_damage'] = data['roof_substrate_damage'].fillna(0)
data['wall_cladding_damage'] = data['wall_cladding_damage'].fillna(0)
data['wall_substrate_damage'] = data['wall_substrate_damage'].fillna(0)
data['wall_structure_damage'] = data['wall_structure_damage'].fillna(0)


# Hurricane Ian design wind speed column is in mph (need to update code developing WiSPD)
conversion_factor = 0.44704
data.loc[data['_project'] == 'Hurricane Ian', 'design_wind_speed'] *= conversion_factor


# WiSPD edits
data['address_municipality'] = data['address_municipality'].str.title()


# Drop all rows that are not wood, single-family structures
fragilities = fragilities[fragilities['ID'].notna()]
fragilities = fragilities[fragilities['ID'].str.contains('W.SF')]
selected_columns = ['ID'] + [f"LS{ls}-Family" for ls in range(1, (NUM_LIMIT_STATES+1))] + \
                    [f"LS{ls}-Theta_0" for ls in range(1, (NUM_LIMIT_STATES+1))] + \
                    [f"LS{ls}-Theta_1" for ls in range(1, (NUM_LIMIT_STATES+1))]

fragilities = fragilities[selected_columns]

# Drop 6s
def filter_id_RDA(id_str):
    parts = id_str.split('.')
    return parts[5] != '6s'

# Drop enhanced garage doors
def filter_id_garage(id_str):
    parts = id_str.split('.')
    return parts[7] != 'sup'

# Apply the function to filter the DataFrame
fragilities = fragilities[fragilities['ID'].apply(filter_id_RDA)]
fragilities = fragilities[fragilities['ID'].apply(filter_id_garage)]

# If using the new fragility functions convert normal to lognormal distributions
# Function to convert normal distribution parameters to lognormal
def normal_to_lognormal(mu, sigma):
    """Convert normal distribution (mu, sigma) to lognormal parameters."""
    # Convert to log-space
    theta_0 = np.log(mu) - 0.5 * np.log(1 + (sigma**2 / mu**2))
    theta_1 = np.sqrt(np.log(1 + (sigma**2 / mu**2)))

    # Convert back to mph scale
    theta_0_mph = np.exp(theta_0)
    return theta_0_mph, theta_1

# Convert normal distributions to lognormal where necessary
limit_states = [1, 2, 3, 4]
for index, row in fragilities.iterrows():
    for ls in limit_states:
        if f"LS{ls}-Family" in fragilities.columns and row[f"LS{ls}-Family"] == "normal":
            mu, sigma = row[f"LS{ls}-Theta_0"], row[f"LS{ls}-Theta_1"]
            fragilities.at[index, f"LS{ls}-Theta_0"], fragilities.at[index, f"LS{ls}-Theta_1"] = normal_to_lognormal(mu, sigma)


# Mapping dictionary - use if using the original fragility function database
rename_mapping = {
    'LS1-Theta_0': 'mu_LS1',
    'LS2-Theta_0': 'mu_LS2',
    'LS3-Theta_0': 'mu_LS3',
    'LS4-Theta_0': 'mu_LS4',
    'LS1-Theta_1': 'log_std_LS1',
    'LS2-Theta_1': 'log_std_LS2',
    'LS3-Theta_1': 'log_std_LS3',
    'LS4-Theta_1': 'log_std_LS4'
}

# Renaming the columns
fragilities.rename(columns=rename_mapping, inplace=True)

    
###############################################################################
'''
GENERATING RANDOM DATA
'''
synthetic_data = pd.DataFrame(data=None, columns=['building_type', 'wind_speed_sample',
                                                  'material', 'occupancy',
                                                  'number_of_stories', 'roof_shape',
                                                  'RWC', 'RDA', 'garage',
                                                  'shutters', 'terrain', 'damage_state', 
                                                  'roof_cover', 'storm_type', 
                                                  'FFE'])
# Randomly sample a wind speed in mph
synthetic_data['wind_speed_sample'] = np.random.randint(75, 200, 50_000)

# Randomly sample a roof cover material
new_column_values = np.random.choice(['Shingle', 'Metal'], size=len(synthetic_data), p=[0.6, 0.4])
synthetic_data['roof_cover'] = new_column_values

# Randomly sample a roof cover material
new_column_values = np.random.choice(['Hurricane', 'Tornado'], size=len(synthetic_data), p=[0.60, 0.40])
synthetic_data['storm_type'] = new_column_values

# Function to randomly select a building type for each wind speed
def select_building_type(row, building_types_df):
    building_types = building_types_df['ID'].tolist()
    return np.random.choice(building_types)

synthetic_data['building_type'] = synthetic_data.apply(select_building_type, 
                                                          args=(fragilities,), 
                                                          axis=1)

synthetic_data[['material', 'occupancy','number_of_stories', 'roof_shape', 'SWB',
                'RDA', 'RWC', 'garage', 'shutters', 'terrain']] = synthetic_data['building_type'].str.split('.').str[:10].tolist()


###############################################################################
'''
ASSIGN A DAMAGE STATE

Using the building ID, select the correct composite fragility function, calculate
a probability of failure using the mu, sigma, and sampled wind speed, take the
difference between the four limit states, randomly generate a value between 0 
and 1, assign a damage class based on the randomly simulated value and the 
difference of the probabilities
'''
# Function to find matching fragility parameters for each building ID
def find_fragility(row):
    for index, frag_row in fragilities.iterrows():
        if frag_row['ID'] in row['building_type']:
            return frag_row[['mu_LS1', 'log_std_LS1', 'mu_LS2', 'log_std_LS2',
                              'mu_LS3', 'log_std_LS3', 'mu_LS4', 'log_std_LS4']]

# Apply the function to the building dataframe
synthetic_data[['mu_LS1', 'log_std_LS1', 'mu_LS2', 'log_std_LS2',
              'mu_LS3', 'log_std_LS3', 'mu_LS4', 'log_std_LS4']] = synthetic_data.apply(find_fragility, axis=1, result_type='expand')

# Functions to calculate probability of failure for each limit state
def probability_of_failure(mu, sigma, wind_speed):
    """Calculate probability of failure using lognormal distribution.
    
    Args:
        mu (float): Scale parameter
        sigma (float): Shape parameter
        wind_speed (float): Wind speed in mph
    
    Returns:
        float: Probability of failure between 0 and 1
    """
    if not all(isinstance(x, (int, float)) for x in [mu, sigma, wind_speed]):
        raise ValueError("All inputs must be numeric")
    if any(x <= 0 for x in [mu, sigma, wind_speed]):
        raise ValueError("All inputs must be positive")
    
    return lognorm.cdf(wind_speed, s=sigma, scale=mu)

def adjust_mu(mu, limit_state, roof_cover, storm_type):
    """Adjust mu based on roof cover material and storm type.
    
    Args:
        mu (float): Original mu value
        limit_state (int): Limit state (1-4)
        roof_cover (str): Type of roof cover
        storm_type (str): Type of storm
    
    Returns:
        float: Adjusted mu value
    """
    adjusted_mu = mu
        
    # Adjustment for roof cover
    if roof_cover in ROOF_COVER_ADJUSTMENTS:
        for states, factor in ROOF_COVER_ADJUSTMENTS[roof_cover].items():
            if limit_state in states:
                adjusted_mu *= factor

    
    # Adjustment for storm type
    if storm_type in STORM_TYPE_ADJUSTMENTS:
        for states, factor in STORM_TYPE_ADJUSTMENTS[storm_type].items():
            if limit_state in states:
                adjusted_mu *= factor
    
    return adjusted_mu

def calculate_failure_probabilities(row):
    """Calculate failure probabilities for all limit states.
    
    Args:
        row (pd.Series): Row of data containing required fields
    
    Returns:
        dict: Probabilities for each limit state
    """
    probabilities = {}
    # static_counter = getattr(calculate_failure_probabilities, 'counter', 0)
    
    try:
        for i in range(1, NUM_LIMIT_STATES + 1):
            mu = row[f'mu_LS{i}']
            sigma = row[f'log_std_LS{i}']
            wind_speed = row['wind_speed_sample']

            # Adjust mu based on building characteristics
            adjusted_mu = adjust_mu(mu, i, row['roof_cover'], row['storm_type'])
            
            # # Print first 5 iterations
            # if static_counter < 5:
            #     print(f"\nIteration {static_counter + 1}:")
            #     print(f"Roof Cover: {row['roof_cover']}")
            #     print(f"Storm Type: {row['storm_type']}")
            #     print(f"Original mu: {mu:.2f}")
            #     print(f"Adjusted mu: {adjusted_mu:.2f}")
            #     print(f"Adjustment Factor: {adjusted_mu/mu:.3f}")
            #     calculate_failure_probabilities.counter = static_counter + 1
            
            probabilities[f'PF_LS{i}'] = probability_of_failure(adjusted_mu, sigma, wind_speed)
            
    except KeyError as e:
        raise KeyError(f"Missing required column: {e}")
    except Exception as e:
        raise Exception(f"Error processing row: {e}")
        
    return probabilities

# Initialize static counter
calculate_failure_probabilities.counter = 0

# Calculate and validate probabilities
failure_probabilities = synthetic_data.apply(calculate_failure_probabilities, axis=1, result_type='expand')
synthetic_data = pd.concat([synthetic_data, failure_probabilities], axis=1)

# Calculate damage state probabilities with validation
damage_state_cols = [f'P(DS={i})' for i in range(5)]
synthetic_data['P(DS=0)'] = 1 - synthetic_data['PF_LS1']
synthetic_data['P(DS=1)'] = synthetic_data['PF_LS1'] - synthetic_data['PF_LS2']
synthetic_data['P(DS=2)'] = synthetic_data['PF_LS2'] - synthetic_data['PF_LS3']
synthetic_data['P(DS=3)'] = synthetic_data['PF_LS3'] - synthetic_data['PF_LS4']
synthetic_data['P(DS=4)'] = synthetic_data['PF_LS4']

# Validate probabilities sum to 1 (within numerical precision)
row_sums = synthetic_data[damage_state_cols].sum(axis=1)
if not np.allclose(row_sums, 1.0, rtol=1e-5):
    print("Warning: Some probability rows don't sum to 1.0")
    print("Max deviation:", abs(1 - row_sums).max())

# Generate random values and assign damage states
np.random.seed(42)  # Optional: for reproducibility
synthetic_data['rand'] = np.random.rand(len(synthetic_data))

def assign_damage_state(row):
    """Assign damage state based on probabilities and random number.
    
    Args:
        row (pd.Series): Row containing damage state probabilities
    
    Returns:
        int: Assigned damage state (0-4)
    """
    probabilities = [row[col] for col in damage_state_cols]
    cumulative_probs = np.cumsum(probabilities)
    
    # Ensure final cumulative probability is 1.0
    cumulative_probs[-1] = 1.0
    
    return np.searchsorted(cumulative_probs, row['rand'])

synthetic_data['damage_state'] = synthetic_data.apply(assign_damage_state, axis=1)

def stratified_trim(df, target_per_class=2000, stratify_cols=['roof_cover', 'storm_type']):
    sampled_frames = []

    for ds in range(5):  # for damage_state 0 through 4
        group = df[df['damage_state'] == ds]

        # Get relative group sizes for stratification
        strata = group.groupby(stratify_cols)
        strata_sizes = strata.size()
        strata_weights = strata_sizes / strata_sizes.sum()

        # How many samples to take from each stratum
        samples_per_stratum = (strata_weights * target_per_class).round().astype(int)

        # Cap at available data to avoid ValueError
        stratum_samples = []
        for index, n_samples in samples_per_stratum.items():
            stratum_df = group
            for col, val in zip(stratify_cols, index if isinstance(index, tuple) else [index]):
                stratum_df = stratum_df[stratum_df[col] == val]

            if len(stratum_df) == 0:
                continue

            take = min(len(stratum_df), n_samples)
            stratum_samples.append(stratum_df.sample(n=take, random_state=42))

        # Combine samples for this damage state
        sampled_frames.append(pd.concat(stratum_samples))

    # Combine all DS groups
    final_df = pd.concat(sampled_frames, ignore_index=True)

    return final_df

# Apply stratified trim to your full synthetic data
trimmed_data = stratified_trim(synthetic_data, target_per_class=2000, stratify_cols=[
    'number_of_stories', 'roof_shape', 'RWC', 'RDA', 'garage', 'shutters', 
    'terrain', 'roof_cover', 'storm_type',
])

# Dictionary to clean/rename column names and values for subplot labeling
clean_labels = {
    "number_of_stories": "Number of Stories",
    "roof_shape": "Roof Shape",
    "RWC": "Roof Wall Connection",
    "RDA": "Roof Deck Attachment",
    "garage": "Garage",
    "shutters": "Shutters",
    "terrain": "Terrain",
    "roof_cover": "Roof Cover",
    "storm_type": "Storm Type",
    # For value cleaning
    "values": {
        "roof_shape": {
            "gab": "Gable",
            "hip": "Hip"
        },
        "shutters": {
            "0": "No shutters",
            "1": "Shutters"
        },
        "RWC": {
            "tnail": "Toe-nail",
            "strap": "Strap"
        },
        "RDA": {
            "6d": "6d 6/12",
            "8d": "8d 6/12",
            "8s": "8d 6/6"
        },
        "garage": {
            "wkd": "Weak",
            "std": "Standard",
            "no": "None"
        }
        # Add more mappings as needed
    }
}

# Create a 3x3 subplot (8 stratify_cols and 1 for damage_state)
fig, axs = plt.subplots(3, 3, figsize=(12,12))
axs = axs.flatten()  # Flatten the 2D array of axes for easy indexing

# List of columns to plot
plot_cols = ['number_of_stories', 'roof_shape', 'RWC', 'RDA', 'garage', 'shutters', 'terrain', 'roof_cover', 'storm_type']

for idx, col in enumerate(plot_cols):
    # Optionally clean values if mapping available
    vc = trimmed_data[col].value_counts().sort_index()
    # If we have value mappings for this col, relabel
    if "values" in clean_labels and col in clean_labels["values"]:
        value_map = clean_labels["values"][col]
        # Convert index using the mapping, fallback to the original if missing
        new_index = [value_map.get(v, v) for v in vc.index]
        vc.index = new_index

    # Bar plot
    vc.plot(
        kind='bar',
        ax=axs[idx],
        color='skyblue',
        edgecolor='black'
    )
    # Use clean label for plot title
    label_name = clean_labels.get(col, col)
    axs[idx].set_title(f'{label_name}', fontsize=12)
    axs[idx].set_ylabel('Count', fontsize=10)
    axs[idx].tick_params(axis='x', rotation=0)

plt.tight_layout() 
plt.savefig(os.path.join(outputPath, 'SyntheticData_Distribution.tiff'), dpi=300)


# Print unique damage states
print("\nUnique damage states and counts:")
print(trimmed_data['damage_state'].value_counts().sort_index())

trimmed_data.to_csv(os.path.join(outputPath, HAZUS_filename))
trimmed_data.to_csv(os.path.join(BOLRPath, HAZUS_filename))


###############################################################################
'''
CREATE NEW DAMAGE STATE COLUMN BASED ON HAZUS FRAGILITY DESCRIPTIONS 
'''
# Create a damage state column based on the HAZUS fragility functions
def determine_damage_state(
    roof_cover_damage, fenestration_avg, roof_substrate_damage,
    roof_structure_damage, wall_cladding_damage, wall_substrate_damage,
    wall_structure_damage, large_door_failure
):
    """
    Determine the damage state based on input attributes.
    
    Parameters:
    - roof_cover_damage: Percentage of roof cover damage (numeric)
    - fenestration_avg: Average number of window/door failures (numeric)
    - roof_substrate_damage: Number of roof panels damaged (numeric)
    - roof_structure_damage: Extent of structural roof damage
    - wall_cladding_damage: Extent of wall cladding damage
    - wall_substrate_damage: Extent of wall substrate damage
    - wall_structure_damage: Extent of wall structure damage
    - large_door_failure: Extent of door damage
    """
    # Convert to boolean internally (any value > 0 indicates damage)
    roof_structure_bool = 1 if roof_structure_damage > 0 else 0
    wall_cladding_bool = 1 if wall_cladding_damage > 0 else 0
    wall_substrate_bool = 1 if wall_substrate_damage > 0 else 0
    wall_structure_bool = 1 if wall_structure_damage > 0 else 0
    large_door_bool = 1 if large_door_failure == 'Yes' else 0

    # Damage State 4: Destruction (check first since most severe)
    if (fenestration_avg > 50 or
        roof_substrate_damage > 25 or
        roof_structure_bool == 1 or 
        wall_structure_bool == 1):
        return 4
    
    # Damage State 0: No damage or very minor damage
    if (roof_cover_damage == 0 and
        fenestration_avg == 0 and
        roof_substrate_damage == 0 and
        roof_structure_bool == 0 and
        wall_cladding_bool == 0 and
        wall_substrate_bool == 0 and
        wall_structure_bool == 0 and
        large_door_bool == 0):
        return 0

    # Damage State 3: Severe damage
    if ((roof_cover_damage > 50 or
         (20 < fenestration_avg <= 50) or
         (10 < roof_substrate_damage <= 25) or
         wall_substrate_bool == 1 or
         wall_cladding_damage > 25) and
        roof_structure_bool == 0 and
        wall_structure_bool == 0):
        return 3

    # Damage State 2: Moderate damage  
    if ((15 < roof_cover_damage <= 50 or
         (5 < fenestration_avg <= 20) or
         (0 < roof_substrate_damage <= 10) or
         (10 < wall_cladding_damage <= 25)) and
        roof_structure_bool == 0 and
        wall_structure_bool == 0):
        return 2

    # Damage State 1: Minor damage
    if ((2 < roof_cover_damage <= 15 or
         fenestration_avg <= 5 or
         large_door_bool == 1 or
         (0 < wall_cladding_damage <= 10)) and
        roof_structure_bool == 0 and
        wall_structure_bool == 0):
        return 1

    # Default: No valid damage state
    return -1

# Apply the function to each row of the dataframe using apply
data['damage_state'] = data.apply(lambda row: determine_damage_state(row['roof_cover_damage'], row['fenestration_avg'], row['roof_substrate_damage'], row['roof_structure_damage'], 
                                                               row['wall_cladding_damage'], row['wall_substrate_damage'], row['wall_structure_damage'], row['large_door_failure']), axis=1)

# Print unique damage states
print("\nUnique damage states and counts:")
print(data['damage_state'].value_counts().sort_index())

# Calculate absolute difference between ratings
data['rating_difference'] = abs(data['wind_damage_rating'] - data['damage_state'])

# Create scatter plot comparing wind damage rating vs damage state
plt.figure(figsize=(10, 6))

# Count number of records at each combination of ratings
rating_counts = data.groupby(['wind_damage_rating', 'damage_state']).size().reset_index(name='count')

# Create scatter plot with count-based colors and sizes
scatter = plt.scatter(rating_counts['wind_damage_rating'], rating_counts['damage_state'],
                     c=rating_counts['count'],
                     s=rating_counts['count']*50, # Scale size based on count
                     alpha=0.6,
                     cmap='YlOrRd') # Yellow to red colormap

plt.xlabel('Wind Damage Rating')
plt.ylabel('HAZUS Damage State')
plt.title('Comparison of Wind Damage Rating vs HAZUS Damage State\nPoint size/color shows number of records')

# Add a perfect correlation line for reference
min_val = min(data['wind_damage_rating'].min(), data['damage_state'].min())
max_val = max(data['wind_damage_rating'].max(), data['damage_state'].max())
plt.plot([min_val, max_val], [min_val, max_val], 'k--', label='Perfect Correlation')

plt.grid(True)
plt.colorbar(scatter, label='Number of Records')
plt.legend()
plt.show()


###############################################################################
'''
GENERATE MISSING WiSPD DATA
'''
# Create a column for structures with design wind speeds >= 130mph
data['basic_wind_speed_mph'] = data['design_wind_speed']*2.237
threshold_ms = 130
data['engineered_str'] = -1
data['engineered_str'] = np.where(data['basic_wind_speed_mph'] >= threshold_ms, 1, 0)

# Modified condition to handle different variations of "None" and 0
data['shutters'] = np.where(
    (data['fenestration_protection'].astype(str).str.lower().isin(['none', '0', 'nan'])) |
    (data['fenestration_protection'] == 0) |
    (pd.isna(data['fenestration_protection'])),
    0, 1)

# Create a garage door type column
def determine_garage(row):
    if row['large_door_present'] == 'yes':
        if row['basic_wind_speed_mph'] < 130:
            return 'wkd'
        elif row['basic_wind_speed_mph'] >= 130:
            return 'std'

    return 'no'  # Return None or another value if conditions don't meet

# Apply the function to each row in the DataFrame
data['garage'] = data.apply(determine_garage, axis=1)

# Create terrain column based on mean_z0
bins = [0, 0.03, 0.15, 0.35, 0.70, 1.00]
labels = [0.03, 0.15, 0.35, 0.70, 1.00]
data['terrain'] = pd.cut(data[terrain_col], bins=bins, labels=labels, include_lowest=True)

# Map building code history to WiSPD dataframe
def get_building_code_values(row, building_codes):
    """
    Get RWC and RDA values based on building code requirements for given location and year.
    Falls back to 'Default' sheet if no match is found and state is not Florida.
    """
    # Handle missing municipality values
    city = str(row['address_municipality']) if pd.notna(row['address_municipality']) else ''
    city = city.strip()
    
    state = row['address_state']
    year_built = row['year_built']
    
    # Convert state abbreviation to full name if needed
    state_full = {'FL': 'Florida'}.get(state, state)
    
    # Create location strings to check against building codes
    location_keys_raw = ([f"{city} {state}", f"{city} {state_full}"] if city else []) + [state, state_full]
    location_keys = [str(k).strip() for k in location_keys_raw if pd.notna(k)]

    # Sanitize sheet name
    sheet_name = row.get('sheet_name', '')
    if not isinstance(sheet_name, str):
        sheet_name = str(sheet_name)
    
    # Step 1: Try to find a matching sheet based on location
    matched_sheet = None
    for candidate in building_codes.sheetnames:
        if any(key in candidate for key in location_keys):
            matched_sheet = candidate
            break
    
    # Step 2: If no match and not Florida, use 'Default'
    if not matched_sheet and state_full != 'Florida':
        if 'Default' in building_codes.sheetnames:
            matched_sheet = 'Default'
    
    # Step 3: If still no match and state is Florida, return hardcoded fallback
    if not matched_sheet:
        return {
            'RWC': 'tnail',
            'RDA': '8d'
        }

    # Step 4: Search code values in the selected sheet
    ws = building_codes[matched_sheet]
    for code_row in ws.iter_rows(min_row=2, values_only=True):
        code, year_adopted, year_ended, rwc, shutters, rda = code_row

        if year_adopted <= year_built <= year_ended:
            return {
                'RWC': 'strap' if rwc == 1 else 'tnail',
                'RDA': rda
            }

    # Step 5: No matching year range found
    return {
        'RWC': 'tnail',
        'RDA': '8d'
    }

# Clean municipality names
data['address_municipality'] = data['address_municipality'].str.strip()
data['address_municipality'] = data['address_municipality'].str.replace(r'\bCounty\b', '', regex=True)

# Convert year_built to numeric, handling invalid values
data['year_built'] = pd.to_numeric(data['year_built'], errors='coerce')

# Apply building code requirements
code_values = data.apply(lambda row: get_building_code_values(row, building_codes), axis=1)
data['RWC'] = code_values.apply(lambda x: x['RWC'])
data['RDA'] = code_values.apply(lambda x: x['RDA'])

# Hurricane Laura update:
data.loc[(data['_project'] == 'Hurricane Laura (2020)') & (data['year_built'] >= 2007), 'RWC'] = 'strap'
data.loc[(data['_project'] == 'Hurricane Laura (2020)') & (data['year_built'] < 2007), 'RWC'] = 'tnail'


# data.to_csv(os.path.join(outputPath, WiSPD_filename))
# data.to_csv(os.path.join(BOLRPath, WiSPD_filename))

###############################################################################
# DAMAGE STATE COMPARISONS
# Create a bar plot of damage_state (synthetic), damage_state (field), and wind_damage_rating (field) on one plot

# Get counts (align indexes to ensure consistent bars for all categories)
synthetic_counts = trimmed_data['damage_state'].value_counts().sort_index()
WiSPD_DS = data['damage_state'].value_counts().sort_index()
WiSPD_WDR = data['wind_damage_rating'].value_counts().sort_index()

# Get all unique states across these three variables (as strings for consistent x-axis tick labels)
all_states = sorted(set(synthetic_counts.index).union(set(WiSPD_DS.index)).union(set(WiSPD_WDR.index)))

# Reindex to ensure all series have the same index/order and fill with 0 if missing
synthetic_bar = synthetic_counts.reindex(all_states, fill_value=0)
field_bar_ds = WiSPD_DS.reindex(all_states, fill_value=0)
field_bar_wdr = WiSPD_WDR.reindex(all_states, fill_value=0)

fig, ax = plt.subplots(figsize=(10, 6))

bar_width = 0.25
index = range(len(all_states))

# Define damage state color mapping (assume categories include only 0 to 4, otherwise fallback to default)
damage_state_colors = {
    0: '#2066ad',      # blue
    1: '#2ca02c',      # green
    2: '#ffce34',      # yellow
    3: '#ff9900',      # orange
    4: '#de2d26'       # red
}

# Convert all_states to integers if possible for correct color selection
def color_for(state):
    try:
        return damage_state_colors[int(state)]
    except Exception:
        return 'grey'

bar_colors = [color_for(s) for s in all_states]

# Plot all three with slight offsets, using correct color mapping for each category
ax.bar([i - bar_width for i in index], synthetic_bar, bar_width, label='Synthetic Data', color=bar_colors, edgecolor='black')
ax.bar(index, field_bar_ds, bar_width, label='WiSPD Data - Damage State', color=bar_colors, edgecolor='black', hatch='///')
ax.bar([i + bar_width for i in index], field_bar_wdr, bar_width, label='WiSPD Data - Wind Damage Rating', color=bar_colors, edgecolor='black', alpha=0.7, hatch='///')

ax.set_xlabel('Damage State', fontsize=12)
ax.set_ylabel('Count', fontsize=12)
ax.set_xticks(list(index))
ax.set_xticklabels(all_states, fontsize=12)

# Move the legend below the x-axis in one row
ax.legend(fontsize=12, loc='upper center', bbox_to_anchor=(0.5, -0.13), ncol=3, frameon=False)

# Set tick label font size for y-axis
ax.tick_params(axis='y', labelsize=12)

plt.tight_layout()
plt.savefig(os.path.join(outputPath, 'DamageState_Comparison_ThreeBars.tiff'), dpi=300)
plt.show()
